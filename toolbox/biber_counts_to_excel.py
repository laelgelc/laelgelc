#!/usr/bin/env python3
"""
Convert fixed-format Biber Tag Count output files to Excel workbooks.

This programme is a Python replacement for the historical SAS-to-Excel shell
workflow used with fixed-format Biber Tag Count output. It reads Biber count
records stored as 12-line fixed-width blocks, extracts Biber's linguistic
feature counts and dimension scores in the established column order, and writes
real Excel `.xlsx` workbooks.

The Biber Tag Count format is assumed to be stable and fixed. Therefore, the
column header definitions are hardcoded and no external header file is required.

Usage examples
--------------

Convert one Biber Tag Count file to one Excel workbook:

    python biber_counts_to_excel.py --input-file counts.txt --output counts.xlsx

Use the default output name for a single input file:

    python biber_counts_to_excel.py --input-file counts.txt

Convert every regular file in a directory to one Excel workbook per input file:

    python biber_counts_to_excel.py --input-dir counts --output-dir excel

Convert every regular file in a directory into one combined Excel workbook:

    python biber_counts_to_excel.py --input-dir counts --combined-output counts.xlsx

Useful options:

    --sheet-name NAME   Worksheet name. Default: biber_counts
    --overwrite         Replace existing output files.
    --verbose           Print progress information.

Notes
-----

- `--input-file` and `--input-dir` are mutually exclusive.
- In single-file mode, the default output path is the input filename with
  `.xlsx` as suffix.
- In directory mode, the default output directory is `excel`.
- Hidden files are ignored in directory mode.
- Excel cells are written as numeric values where possible, except for the
  `filename` column, which is always text.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BIBER_COLUMNS = [
    "filename",
    "ttr",
    "wrlengh",
    "wcount",
    "prv_vb",
    "that_del",
    "contrac",
    "pres",
    "pro2",
    "pro_do",
    "pdem",
    "gen_emph",
    "pro1",
    "it",
    "be_state",
    "sub_cos",
    "prtcle",
    "pany",
    "gen_hdg",
    "amplifr",
    "wh_ques",
    "pos_mod",
    "o_and",
    "wh_cl",
    "finlprep",
    "n",
    "prep",
    "adj_attr",
    "pasttnse",
    "pro3",
    "perfects",
    "pub_vb",
    "rel_obj",
    "rel_subj",
    "rel_pipe",
    "p_and",
    "n_nom",
    "tm_adv",
    "pl_adv",
    "advs",
    "inf",
    "prd_mod",
    "sua_vb",
    "sub_cnd",
    "nec_mod",
    "spl_aux",
    "conjncts",
    "agls_psv",
    "by_pasv",
    "whiz_vbn",
    "sub_othr",
    "vcmo",
    "downtone",
    "pred_adj",
    "allmodal",
    "allconj",
    "allpasv",
    "allwh",
    "allwhrel",
    "alladj",
    "allpro",
    "have",
    "allverb",
    "vprogrsv",
    "that_rel",
    "jcmp",
    "nonf_vth",
    "att_vth",
    "fact_vth",
    "lkly_vth",
    "att_jth",
    "fact_jth",
    "lkly_jth",
    "nfct_nth",
    "att_nth",
    "fct_nth",
    "lkly_nth",
    "spch_vto",
    "mntl_vto",
    "dsre_vto",
    "efrt_vto",
    "prob_vto",
    "x1_jto",
    "x2_jto",
    "x3_jto",
    "x4_jto",
    "x5_jto",
    "all_nto",
    "nonfadvl",
    "atadvl",
    "factadvl",
    "lklydvl",
    "all_vth",
    "all_jth",
    "all_nth",
    "all_th",
    "all_vto",
    "all_jto",
    "all_to",
    "all_advl",
    "act_ipv",
    "act_tpv",
    "mentalpv",
    "commpv",
    "occurpv",
    "copulapv",
    "aspectpv",
    "humann",
    "prcessn",
    "cognitn",
    "abstrcn",
    "concrtn",
    "tccncrt",
    "quann",
    "placen",
    "groupn",
    "sizej",
    "timej",
    "colorj",
    "evalj",
    "relatnj",
    "topicj",
    "actv",
    "commv",
    "mentalv",
    "causev",
    "occurv",
    "existv",
    "aspectv",
    "dim1",
    "dim2",
    "dim3",
    "dim4",
    "dim5",
]


LEGACY_OUTPUT_ORDER = [
    "filename",
    "c11",
    "c12",
    "c13",
    "c21",
    "c22",
    "c23",
    "c24",
    "c25",
    "c26",
    "c27",
    "c28",
    "c29",
    "c210",
    "c211",
    "c212",
    "c213",
    "c214",
    "c215",
    "c31",
    "c32",
    "c33",
    "c34",
    "c35",
    "c36",
    "c37",
    "c38",
    "c39",
    "c310",
    "c311",
    "c312",
    "c313",
    "c314",
    "c315",
    "c41",
    "c42",
    "c43",
    "c44",
    "c45",
    "c46",
    "c47",
    "c48",
    "c49",
    "c410",
    "c411",
    "c412",
    "c413",
    "c414",
    "c415",
    "c51",
    "c52",
    "c53",
    "c54",
    "c55",
    "c56",
    "c57",
    "c58",
    "c59",
    "c510",
    "c511",
    "c512",
    "c513",
    "c514",
    "c515",
    "c61",
    "c62",
    "c74",
    "c75",
    "c76",
    "c77",
    "c78",
    "c79",
    "c710",
    "c711",
    "c712",
    "c713",
    "c714",
    "c715",
    "c81",
    "c82",
    "c83",
    "c84",
    "c85",
    "c86",
    "c87",
    "c88",
    "c89",
    "c810",
    "c811",
    "c812",
    "c813",
    "c814",
    "c91",
    "c92",
    "c93",
    "c94",
    "c95",
    "c96",
    "c97",
    "c98",
    "c101",
    "c102",
    "c103",
    "c104",
    "c105",
    "c106",
    "c107",
    "c108",
    "c109",
    "c1010",
    "c1011",
    "c1012",
    "c1013",
    "c1014",
    "c1015",
    "c111",
    "c112",
    "c113",
    "c114",
    "c115",
    "c116",
    "c117",
    "c118",
    "c119",
    "c1110",
    "c1111",
    "c1112",
    "c1113",
    "c1114",
    "c121",
    "c122",
    "c123",
    "c124",
    "c125",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert fixed-format Biber Tag Count output to Excel .xlsx."
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--input-file",
        type=Path,
        help="Single Biber Tag Count file to parse.",
    )
    input_group.add_argument(
        "--input-dir",
        type=Path,
        help="Directory containing Biber Tag Count files to parse.",
    )

    parser.add_argument(
        "--output",
        type=Path,
        help="Output .xlsx path for single-file mode.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Output directory for per-file .xlsx files in directory mode.",
    )
    parser.add_argument(
        "--combined-output",
        type=Path,
        help="Output .xlsx path for one combined workbook in directory mode.",
    )
    parser.add_argument(
        "--sheet-name",
        default="biber_counts",
        help="Excel worksheet name. Default: biber_counts.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing output files.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress information.",
    )

    return parser.parse_args()


def fixed_slice(line: str, start: int, end: int | None = None) -> str:
    """Return a 1-based fixed-width substring with spaces removed."""
    start_index = start - 1

    if end is None:
        return line[start_index:].strip().replace(" ", "")

    return line[start_index:end].strip().replace(" ", "")


def coerce_value(value: str) -> str | int | float:
    """Convert a parsed field to int or float when possible."""
    value = value.strip()

    if value == "":
        return ""

    try:
        number = float(value)
    except ValueError:
        return value

    if number.is_integer() and "." not in value:
        return int(number)

    return number


def normalise_lines(text: str) -> list[str]:
    """Remove carriage returns and split text into lines."""
    return text.replace("\r", "").splitlines()


def chunk_records(lines: list[str], record_size: int = 12) -> list[list[str]]:
    """Split input lines into non-empty fixed-size Biber records."""
    records: list[list[str]] = []

    for start in range(0, len(lines), record_size):
        record = lines[start:start + record_size]

        if not any(line.strip() for line in record):
            continue

        if len(record) != record_size:
            raise ValueError(
                f"Incomplete Biber record starting at line {start + 1}: "
                f"expected {record_size} lines, got {len(record)}"
            )

        records.append(record)

    return records


def parse_record(record: list[str]) -> dict[str, str]:
    """
    Parse one 12-line Biber fixed-format record into legacy c-field names.

    Line 1 contains filename, type/token ratio, word length and word count.
    Lines 2-11 contain 15 fixed-width fields of five characters each.
    Line 12 contains five fixed-width fields of ten characters each.
    """
    if len(record) != 12:
        raise ValueError(f"Expected 12 lines per record, got {len(record)}")

    fields: dict[str, str] = {}

    first_line = record[0]
    fields["filename"] = fixed_slice(first_line, 1, 60)
    fields["c11"] = fixed_slice(first_line, 61, 65)
    fields["c12"] = fixed_slice(first_line, 66, 70)
    fields["c13"] = fixed_slice(first_line, 71, None)

    for line_number in range(2, 12):
        line = record[line_number - 1]

        for field_number in range(1, 16):
            start = ((field_number - 1) * 5) + 1
            end = field_number * 5
            fields[f"c{line_number}{field_number}"] = fixed_slice(line, start, end)

    twelfth_line = record[11]

    for field_number in range(1, 6):
        start = ((field_number - 1) * 10) + 1
        end = field_number * 10
        fields[f"c12{field_number}"] = fixed_slice(twelfth_line, start, end)

    return fields


def record_to_row(record: list[str]) -> list[Any]:
    """Convert one fixed-format Biber record to one Excel row."""
    fields = parse_record(record)
    raw_row = [fields.get(field_name, "") for field_name in LEGACY_OUTPUT_ORDER]

    if len(raw_row) != len(BIBER_COLUMNS):
        raise ValueError(
            f"Column mismatch: parsed {len(raw_row)} values, "
            f"expected {len(BIBER_COLUMNS)}"
        )

    return [
        value if column_name == "filename" else coerce_value(value)
        for column_name, value in zip(BIBER_COLUMNS, raw_row)
    ]


def parse_biber_file(input_path: Path) -> list[list[Any]]:
    """Parse a Biber Tag Count file into Excel-ready rows."""
    text = input_path.read_text(encoding="utf-8", errors="replace")
    lines = normalise_lines(text)
    records = chunk_records(lines)
    rows = [record_to_row(record) for record in records]

    for row_number, row in enumerate(rows, start=1):
        if len(row) != len(BIBER_COLUMNS):
            raise ValueError(
                f"{input_path}: row {row_number} has {len(row)} values, "
                f"expected {len(BIBER_COLUMNS)}"
            )

    return rows


def safe_sheet_name(sheet_name: str) -> str:
    """Return a valid Excel worksheet name."""
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    cleaned = sheet_name

    for char in invalid_chars:
        cleaned = cleaned.replace(char, "_")

    cleaned = cleaned.strip() or "biber_counts"

    return cleaned[:31]


def write_xlsx(
        output_path: Path,
        columns: list[str],
        rows: list[list[Any]],
        sheet_name: str = "biber_counts",
        overwrite: bool = False,
) -> None:
    """Write rows to a real Excel .xlsx workbook."""
    if output_path.exists() and not overwrite:
        raise FileExistsError(
            f"Output file already exists: {output_path}. "
            "Use --overwrite to replace it."
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_sheet_name(sheet_name)

    worksheet.append(columns)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(columns, start=1):
        column_letter = get_column_letter(column_index)

        if column_name == "filename":
            width = 32
        else:
            width = max(10, min(len(column_name) + 2, 18))

        worksheet.column_dimensions[column_letter].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def default_single_output_path(input_path: Path) -> Path:
    """Return the default output path for a single input file."""
    return input_path.with_suffix(".xlsx")


def output_path_for_input(input_path: Path, output_dir: Path) -> Path:
    """Return the per-file output path for directory mode."""
    return output_dir / f"{input_path.stem}.xlsx"


def discover_input_files(input_dir: Path) -> list[Path]:
    """Return non-hidden regular files in a directory."""
    return sorted(
        path
        for path in input_dir.iterdir()
        if path.is_file() and not path.name.startswith(".")
    )


def run_single_file_mode(args: argparse.Namespace) -> None:
    input_path: Path = args.input_file

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if not input_path.is_file():
        raise ValueError(f"Input path is not a file: {input_path}")

    if args.combined_output is not None:
        raise ValueError("--combined-output is only supported with --input-dir")

    output_path = args.output or default_single_output_path(input_path)

    if args.output_dir is not None:
        raise ValueError("--output-dir is only supported with --input-dir")

    if args.verbose:
        print(f"Parsing {input_path}")

    rows = parse_biber_file(input_path)

    if args.verbose:
        print(f"Parsed {len(rows)} records")
        print(f"Writing {output_path}")

    write_xlsx(
        output_path=output_path,
        columns=BIBER_COLUMNS,
        rows=rows,
        sheet_name=args.sheet_name,
        overwrite=args.overwrite,
    )


def run_directory_mode(args: argparse.Namespace) -> None:
    input_dir: Path = args.input_dir

    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    if not input_dir.is_dir():
        raise ValueError(f"Input path is not a directory: {input_dir}")

    if args.output is not None:
        raise ValueError("--output is only supported with --input-file")

    input_files = discover_input_files(input_dir)

    if not input_files:
        raise ValueError(f"No regular input files found in: {input_dir}")

    if args.combined_output is not None:
        combined_rows: list[list[Any]] = []

        for input_path in input_files:
            if args.verbose:
                print(f"Parsing {input_path}")

            rows = parse_biber_file(input_path)
            combined_rows.extend(rows)

            if args.verbose:
                print(f"Parsed {len(rows)} records from {input_path.name}")

        if args.verbose:
            print(f"Writing combined workbook {args.combined_output}")
            print(f"Total records: {len(combined_rows)}")

        write_xlsx(
            output_path=args.combined_output,
            columns=BIBER_COLUMNS,
            rows=combined_rows,
            sheet_name=args.sheet_name,
            overwrite=args.overwrite,
        )

        return

    output_dir = args.output_dir or Path("excel")

    for input_path in input_files:
        output_path = output_path_for_input(input_path, output_dir)

        if args.verbose:
            print(f"Parsing {input_path}")

        rows = parse_biber_file(input_path)

        if args.verbose:
            print(f"Parsed {len(rows)} records")
            print(f"Writing {output_path}")

        write_xlsx(
            output_path=output_path,
            columns=BIBER_COLUMNS,
            rows=rows,
            sheet_name=args.sheet_name,
            overwrite=args.overwrite,
        )


def main() -> int:
    args = parse_args()

    try:
        if args.input_file is not None:
            run_single_file_mode(args)
        else:
            run_directory_mode(args)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    if args.verbose:
        print("Done.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())